from pyhunter import PyHunter
from openpyxl import Workbook
import getpass


def Busqueda(organizacion):
    # Cantidad de resultados esperados de la búsqueda
    # El límite MENSUAL de Hunter es 50, cuidado!
    resultado = hunter.domain_search(company=organizacion, limit=1,
                                     emails_type='personal')
    return resultado


def GuardarInformacion(datosEncontrados, organizacion):
    libro = Workbook()
    hoja = libro.create_sheet(organizacion)
    libro.save("Hunter" + organizacion + ".xlsx")
    hoja["A1"] = "nombre"
    hoja['B1'] = 'correo '
    lista = datosEncontrados.get("emails")
    cont = 2
    cont2 = 1
    for i in range(len(lista)):
        hoja.cell(cont, cont2 + 1, lista[i]['value'])
        hoja.cell(cont, cont2, lista[i]['first_name'] + ' ' +
                  str(lista[i]["last_name"]))
        cont += 1

    print(lista)

    libro.save("Hunter" + organizacion + ".xlsx")
print("Script para buscar información")
apikey = getpass.getpass("Ingresa tu API:")
hunter = PyHunter(apikey)
orga = input("Dominio a investigar: ")


datosEncontrados = Busqueda(orga)
if datosEncontrados is None:
    exit()
else:
    print(datosEncontrados)
    print(type(datosEncontrados))
    GuardarInformacion(datosEncontrados, orga)